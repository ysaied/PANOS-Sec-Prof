#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Panorama security-rules extractor
"""
#pip3 install requests xmltodict openpyxl

import sys
import logging
from datetime import datetime

from panos_actions import panos_op_cmd, panos_config_show  # Import API helper functions
from openpyxl import Workbook
from openpyxl.styles import Font

# Import API key from a separate credentials module or environment variable
try:
    from credentials import api_key  # credentials.py should define the variable `api_key`
except ImportError:
    print("Error: Unable to import API key from credentials. Ensure credentials.py exists with api_key defined.")
    sys.exit(1)

# Panorama connection settings
PANORAMA_IP = "192.168.2.31"  # Update this to your Panorama's IP or hostname
OUTPUT_FILE = "Panorama_Security_Rules.xlsx"

# Configure logging to file and console with timestamp, level, and message.
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("panorama_rules.log", mode='w'),
        logging.StreamHandler(sys.stdout)
    ]
)

def get_device_groups(panorama_ip, api_key):
    """
    Retrieve all device groups from Panorama.
    Returns a list of device group names.
    """
    logging.info("Retrieving device groups from Panorama %s...", panorama_ip)
    xml_cmd = "<show><devicegroups></devicegroups></show>"
    device_groups = []
    try:
        response = panos_op_cmd(fw_ip=panorama_ip, fw_key=api_key, xml_cmd=xml_cmd)
    except Exception as e:
        logging.error("Failed to retrieve device groups: %s", e)
        return []
    # Parse the response for device group entries
    result = response.get("result", {})
    dg_section = result.get("devicegroups") if result else None
    if not dg_section:
        logging.warning("No device group information found in Panorama response.")
        return []
    entries = dg_section.get("entry")
    if not entries:
        logging.info("No device groups defined on Panorama.")
        return []
    # Normalize entries to a list
    if isinstance(entries, list):
        for entry in entries:
            name = entry.get("@name")
            if name:
                device_groups.append(name)
    elif isinstance(entries, dict):
        name = entries.get("@name")
        if name:
            device_groups.append(name)
    logging.info("Retrieved %d device groups: %s", len(device_groups), device_groups)
    return device_groups

def get_security_rules_for_group(panorama_ip, api_key, device_group, rulebase_type="pre"):
    """
    Retrieve security policy rules for a given device group and rulebase type ('pre' or 'post').
    Returns a list of rule dictionaries.
    """
    xpath = (f"/config/devices/entry/device-group/entry[@name='{device_group}']"
             f"/{rulebase_type}-rulebase/security")
    try:
        response = panos_config_show(fw_ip=panorama_ip, fw_key=api_key, xpath=xpath)
    except Exception as e:
        logging.error("Error retrieving %s-rulebase for device group '%s': %s", rulebase_type.upper(), device_group, e)
        return []
    # Check for successful response
    if response.get("@status") != "success":
        err_msg = response.get("msg", {}).get("line", "Unknown error")
        logging.error("Failed to get %s-rulebase for %s (API status: %s, message: %s)",
                      rulebase_type.upper(), device_group, response.get("@status"), err_msg)
        return []
    result = response.get("result", {})
    if not result or "security" not in result:
        # No security rules section found (possibly no rules configured)
        return []
    security = result["security"]
    if "rules" not in security or not security["rules"]:
        # Security policy container exists but no rules inside
        return []
    rules_container = security["rules"]
    entries = rules_container.get("entry")
    if not entries:
        return []
    # Normalize to list of rule entries
    if isinstance(entries, list):
        return entries
    elif isinstance(entries, dict):
        return [entries]
    else:
        logging.warning("Unexpected format in rules for device group '%s' (%s-rulebase).", device_group, rulebase_type)
        return []

def extract_profile_name(rule):
    """
    Extract the Security Profile Group name associated with a rule, if any.
    Returns an empty string if no profile group is attached.
    Note: Rules with individually assigned profiles (not using a group) will be represented as empty here.
    """
    profile_setting = rule.get("profile-setting")
    if not profile_setting:
        return ""
    # Check if a profile group is specified
    group = profile_setting.get("group")
    if group and "member" in group:
        member = group["member"]
        if not member:
            return ""
        if isinstance(member, list):
            # Join multiple profile group names if present (typically only one group is allowed)
            return ", ".join(member)
        else:
            return str(member)
    # If profiles are defined individually (under 'profile-setting' -> 'profiles'), we do not aggregate them here
    return ""

def export_rules_to_excel(rules_data, output_file):
    """
    Export the compiled security rules data to an Excel file.
    rules_data: list of rows (each row is a list or tuple of [Device Group, Pre/Post, Rule Name, Action, Security Profile]).
    """
    logging.info("Exporting %d rules to Excel file '%s'...", len(rules_data), output_file)
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Rules"
    # Define and write header row
    headers = ["Device Group", "Pre/Post", "Rule Name", "Action", "Security Profile"]
    ws.append(headers)
    # Style the header row (bold font)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Write data rows
    for row in rules_data:
        ws.append(list(row))
    # Adjust column widths to fit content
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter  # e.g., 'A', 'B', 'C', etc.
        for cell in col:
            if cell.value is not None:
                # Determine text length of cell value
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        # Add some padding to the width
        ws.column_dimensions[column_letter].width = max_length + 2
    # Freeze the header row and apply an autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    # Save the workbook to a file
    try:
        wb.save(output_file)
        wb.close()
        logging.info("Excel file '%s' successfully saved.", output_file)
    except Exception as e:
        logging.error("Failed to save Excel file '%s': %s", output_file, e)
        raise

def main():
    start_time = datetime.now()
    logging.info("Script started. Fetching device groups and security rules...")

    device_groups = get_device_groups(PANORAMA_IP, api_key)
    if not device_groups:
        logging.error("No device groups retrieved from Panorama. Exiting.")
        return  # Exit early if we cannot get any device group

    all_rules = []  # List to collect all rule data rows
    for dg in device_groups:
        logging.info("Processing device group: %s", dg)
        # Fetch Pre-rulebase security rules
        pre_rules = get_security_rules_for_group(PANORAMA_IP, api_key, dg, rulebase_type="pre")
        if pre_rules:
            logging.info("  Found %d pre-rules in device group '%s'.", len(pre_rules), dg)
            for rule in pre_rules:
                rule_name = rule.get("@name", "<unnamed>")
                action = rule.get("action", "")
                profile = extract_profile_name(rule)
                all_rules.append((dg, "PRE-RULE", rule_name, action, profile))
        else:
            logging.info("  No Pre-Rules in device group '%s'.", dg)
        # Fetch Post-rulebase security rules
        post_rules = get_security_rules_for_group(PANORAMA_IP, api_key, dg, rulebase_type="post")
        if post_rules:
            logging.info("  Found %d post-rules in device group '%s'.", len(post_rules), dg)
            for rule in post_rules:
                rule_name = rule.get("@name", "<unnamed>")
                action = rule.get("action", "")
                profile = extract_profile_name(rule)
                all_rules.append((dg, "POST-RULE", rule_name, action, profile))
        else:
            logging.info("  No Post-Rules in device group '%s'.", dg)

    # Export all collected rules to an Excel spreadsheet
    try:
        export_rules_to_excel(all_rules, OUTPUT_FILE)
    except Exception as e:
        logging.error("Error during Excel export: %s", e)
        sys.exit(1)  # Critical failure writing the file

    end_time = datetime.now()
    elapsed = (end_time - start_time).total_seconds()
    logging.info("Script completed in %.2f seconds. Output file: %s", elapsed, OUTPUT_FILE)

if __name__ == "__main__":
    main()